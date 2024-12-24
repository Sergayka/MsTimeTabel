import os

from openpyxl import load_workbook
import json

# Укажите путь к вашему Excel-файлу
file_path = 'TeachersShedule.xlsx'

# Загружаем рабочую книгу
workbook = load_workbook(filename=file_path)

# Получаем первый лист
first_sheet = workbook.sheetnames[20]
sheet = workbook[first_sheet]

# Читаем данные первых трех столбцов и проверяем объединение ячеек
data = []
# [0] - 18, [1] - 9, [2] - 16, [3] - 9, [4] - 9, [5] - 22, [6] - 21, [7] - 21,
# [8] - 21, [9] - 21, [10] - 21, [11] - 16, [12] - 16, [13] - 21, [14] - 21, [15] -21 ,
# [16] -, [17] -, [18] -, [19] -, [20] -,
for row in sheet.iter_rows(min_col=1, max_col=30, values_only=False):  # max_col теперь max_column
    row_data = []
    for cell in row:
        # Проверяем, объединена ли ячейка
        is_merged = any(cell.coordinate in merged_range for merged_range in sheet.merged_cells)
        row_data.append((cell.value, is_merged))
    data.append(row_data)

# print(data)

# Структура для хранения расписания
schedule = {}


# Функция для добавления урока в расписание
def add_lesson(week, group, day, time_slot, lesson):
    if group not in schedule:
        schedule[group] = {
            "Четная неделя": {},
            "Нечетная неделя": {}
        }
    if day not in schedule[group][week]:
        schedule[group][week][day] = {}
    schedule[group][week][day][time_slot] = lesson


# Переменные для текущей недели и дня
is_even_week = True  # Начинаем с четной недели
current_day = None  # Переменная для хранения текущего дня недели

# Обрабатываем данные для всех групп
for i in range(1, len(data) - 1, 2):
    # Обрабатываем дни недели и время для каждой группы
    day = None  # День недели (столбец A)
    time_slot = data[i][1][0]  # Время занятия (столбец B)

    if data[i][0][0] is not None:
        day = data[i][0][0]

    if day != current_day and day is not None:
        current_day = day  # Обновляем текущий день недели

    # Перебираем все группы, начиная с 3-го столбца
    for group_index in range(2, len(data[i])):
        # print(data[i])
        teacher_fio = data[0][group_index][0]  # Название группы из первой строки (ячейка первого столбца текущей группы)
        # print(group_name)
        # break
        merged = data[i][group_index][1]  # Статус объединенности ячейки
        lesson = data[i][group_index][0]  # Название урока

        # Добавляем урок в расписание
        if merged:
            # Если ячейка объединена, то урок будет и в четной, и в нечетной неделе
            add_lesson("Четная неделя", teacher_fio, current_day, time_slot, lesson)
            add_lesson("Нечетная неделя", teacher_fio, current_day, time_slot, lesson)
        else:
            # Если ячейка не объединена, то одно занятие для четной недели, другое для нечетной
            add_lesson("Четная неделя", teacher_fio, current_day, time_slot, lesson)
            add_lesson("Нечетная неделя", teacher_fio, current_day, time_slot,
                       data[i + 1][group_index][0])  # Следующий урок для нечетной недели

#Создадим папку для сохранения файлов, если она не существует
output_dir = "teachers_schedule_jsons"
os.makedirs(output_dir, exist_ok=True)

# Сохраняем расписание для каждой группы в отдельный JSON файл
for group in schedule:
    group_schedule = schedule[group]
    group_file_path = os.path.join(output_dir, f"{group}_schedule.json")
    with open(group_file_path, "w", encoding="utf-8") as json_file:
        json.dump(group_schedule, json_file, ensure_ascii=False, indent=4)

    print(f"Расписание для группы {group} сохранено в файл {group_file_path}")
